import streamlit as st
import openpyxl
from openpyxl import Workbook
import csv
from docxtpl import DocxTemplate, InlineImage
from docx.shared import Cm
import os
import tempfile
import zipfile
import pandas as pd
import requests
import platform
import subprocess
from pathlib import Path
from requests.auth import HTTPBasicAuth
from io import StringIO, BytesIO

st.set_page_config(page_title="Gerador de Relatórios SEPE", layout="wide")

st.title("🏗️ Gerador de Relatórios de Vistoria")
st.markdown("---")

# ── Tabs de fonte de dados ────────────────────────────────────────────────────
tab1, tab2 = st.tabs(["📡 Conectar ao ODK Central", "📁 Upload de Arquivo CSV"])

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 1 — ODK Central
# ═══════════════════════════════════════════════════════════════════════════════
with tab1:
    st.subheader("🔗 Conexão com ODK Central")

    col_odk1, col_odk2 = st.columns(2)
    with col_odk1:
        odk_url        = st.text_input("URL do ODK Central",
                                       value="https://levantamentos.dflegal.df.gov.br")
        odk_project_id = st.text_input("ID do Projeto", value="4")
        odk_form_id    = st.text_input("ID do Formulário",
                                       value="aWX8oXGiD9zmcpDX7KtAer")
    with col_odk2:
        odk_email     = st.text_input("Email")
        odk_password  = st.text_input("Senha", type="password")
        baixar_anexos = st.checkbox("Baixar anexos (imagens)", value=True)

    if st.button("🔄 Conectar e Buscar Dados", type="primary", use_container_width=True):
        if not odk_email or not odk_password:
            st.error("❌ Por favor, preencha email e senha!")
        else:
            try:
                with st.spinner("Conectando ao ODK Central..."):
                    import io
                    from zipfile import ZipFile

                    base_url = (f"{odk_url}/v1/projects/{odk_project_id}"
                                f"/forms/{odk_form_id}")
                    auth = HTTPBasicAuth(odk_email, odk_password)

                    st.info("Buscando dados do formulário...")
                    response = requests.get(f"{base_url}/submissions.csv.zip",
                                            auth=auth)
                    response.raise_for_status()

                    with ZipFile(io.BytesIO(response.content)) as zf:
                        csv_fn      = [f for f in zf.namelist()
                                       if f.endswith('.csv')][0]
                        csv_content = zf.read(csv_fn).decode('utf-8')

                    st.session_state['csv_data']    = csv_content
                    st.session_state['data_source'] = 'odk'
                    st.session_state['csv_bytes']   = csv_content.encode('utf-8')
                    num_linhas = len(csv_content.split('\n')) - 1

                    # Tentar salvar CSV localmente
                    try:
                        pasta_sepe = Path('C:/sepe')
                        pasta_sepe.mkdir(parents=True, exist_ok=True)
                        csv_local = pasta_sepe / 'dados_odk.csv'
                        csv_local.write_text(csv_content, encoding='utf-8')
                        if csv_local.exists() and csv_local.stat().st_size > 0:
                            st.session_state['csv_pasta_local'] = str(pasta_sepe)
                            st.success(
                                f"💾 CSV salvo localmente: {csv_local} "
                                f"({csv_local.stat().st_size} bytes)")
                    except Exception:
                        st.info("ℹ️ App em servidor remoto — "
                                "use o botão ⬇️ para baixar o CSV.")

                    # ── Download de anexos ────────────────────────────────────
                    if baixar_anexos:
                        st.info("Verificando anexos no servidor ODK...")

                        pasta_media = Path('C:/sepe/media')
                        pasta_media.mkdir(parents=True, exist_ok=True)
                        pasta_temp  = Path(tempfile.gettempdir()) / 'odk_media'
                        pasta_temp.mkdir(parents=True, exist_ok=True)

                        try:
                            arquivos_no_hd = set(os.listdir(str(pasta_media)))
                            st.info(f"📂 {len(arquivos_no_hd)} arquivos já "
                                    f"existem em C:/sepe/media/")

                            # Montar id_map
                            id_map = {}
                            for row in csv.DictReader(io.StringIO(csv_content)):
                                iid_raw = (row.get('KEY')
                                           or row.get('InstanceID')
                                           or row.get('meta-instanceID'))
                                if not iid_raw:
                                    continue
                                id_proj = row.get('details-N_mero_ID')
                                if not id_proj or not str(id_proj).strip():
                                    vals = list(row.values())
                                    id_proj = vals[4] if len(vals) > 4 else None
                                if id_proj:
                                    id_proj = str(id_proj).strip()
                                    alternativas = [iid_raw]
                                    if iid_raw.startswith('uuid:'):
                                        alternativas.append(iid_raw[5:])
                                    else:
                                        alternativas.append(f"uuid:{iid_raw}")
                                    for k in alternativas:
                                        id_map[k] = id_proj

                            st.success(
                                f"✅ {len(set(id_map.values()))} registros mapeados")

                            subs = requests.get(f"{base_url}/submissions",
                                                auth=auth)
                            subs.raise_for_status()

                            todos, para_baixar = [], []
                            st.info("🔍 Comparando com arquivos locais...")

                            for sub in subs.json():
                                iid     = sub.get('instanceId')
                                id_proj = id_map.get(iid)
                                att_url = (f"{base_url}/submissions/"
                                           f"{iid}/attachments")
                                att_r   = requests.get(att_url, auth=auth)
                                if att_r.status_code != 200:
                                    continue
                                for att in att_r.json():
                                    nome = att.get('name')
                                    if not nome:
                                        continue
                                    entrada = {
                                        'nome_original':    nome,
                                        'nome_com_prefixo': (
                                            f"foto_{id_proj}_{nome}"
                                            if id_proj else f"foto_{nome}"),
                                        'attachments_url':  att_url,
                                        'caminho_local':    str(pasta_media / nome),
                                        'caminho_temp':     str(pasta_temp  / nome),
                                    }
                                    todos.append(entrada)
                                    if nome not in arquivos_no_hd:
                                        para_baixar.append(entrada)

                            total    = len(todos)
                            ja_tem   = total - len(para_baixar)
                            faltando = len(para_baixar)
                            baixados_ok = erros = 0

                            st.info(
                                f"📊 Servidor: **{total}** anexos — "
                                f"**{ja_tem}** já no HD, "
                                f"**{faltando}** para baixar")

                            if faltando == 0:
                                st.success("✅ Todos os anexos já estão "
                                           "em C:/sepe/media/")
                            else:
                                prog   = st.progress(0)
                                status = st.empty()
                                for i, e in enumerate(para_baixar, 1):
                                    status.text(f"Baixando {i}/{faltando}: "
                                                f"{e['nome_original']}")
                                    prog.progress(i / faltando)
                                    try:
                                        r = requests.get(
                                            f"{e['attachments_url']}/"
                                            f"{e['nome_original']}",
                                            auth=auth, timeout=30)
                                        if r.status_code == 200:
                                            Path(e['caminho_local']).write_bytes(
                                                r.content)
                                            Path(e['caminho_temp']).write_bytes(
                                                r.content)
                                            baixados_ok += 1
                                        else:
                                            erros += 1
                                    except Exception as ex:
                                        st.warning(f"⚠️ {e['nome_original']}: "
                                                   f"{ex}")
                                        erros += 1
                                prog.empty()
                                status.empty()

                            registrados = []
                            for e in todos:
                                p = Path(e['caminho_local'])
                                if p.exists():
                                    registrados.append({
                                        'nome_original':    e['nome_original'],
                                        'nome_com_prefixo': e['nome_com_prefixo'],
                                        'path_temp':        str(p),
                                        'data':             p.read_bytes(),
                                    })
                            st.session_state['anexos_baixados'] = registrados

                            msg = (f"✅ {total} anexos — {ja_tem} já existiam, "
                                   f"{baixados_ok} baixados")
                            if erros:
                                msg += f", ⚠️ {erros} erros"
                            st.success(msg)

                        except Exception as e:
                            st.warning(f"⚠️ Erro ao baixar anexos: {e}")
                            st.exception(e)

                    st.success(f"✅ Conectado! {num_linhas} registros.")
                    st.rerun()

            except Exception as e:
                st.error(f"❌ Erro ao conectar: {e}")
                st.exception(e)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 2 — Upload CSV + diretório manual de fotos
# ═══════════════════════════════════════════════════════════════════════════════
with tab2:
    st.subheader("📄 Upload Manual de CSV")
    csv_file_upload = st.file_uploader("Selecione o arquivo CSV", type=['csv'])
    if csv_file_upload:
        st.session_state['csv_data']    = csv_file_upload.getvalue().decode('utf-8')
        st.session_state['data_source'] = 'upload'
        st.success("✅ Arquivo CSV carregado com sucesso!")

    st.markdown("---")
    st.subheader("🖼️ Diretório de Fotos")
    st.caption("Informe a pasta onde estão as imagens referenciadas no CSV.")

    dir_fotos_input = st.text_input(
        "Caminho da pasta de fotos",
        value=st.session_state.get('dir_fotos_manual', 'C:/sepe/media'),
        placeholder="Ex: C:/sepe/media  ou  D:/fotos_vistoria",
        help="A pasta deve ter os arquivos com os mesmos nomes do formulário ODK."
    )

    col_dir1, col_dir2 = st.columns(2)
    with col_dir1:
        if st.button("✅ Usar esta pasta", use_container_width=True,
                     key="btn_usar_dir_fotos"):
            if os.path.isdir(dir_fotos_input):
                st.session_state['dir_fotos_manual'] = dir_fotos_input
                n = len([f for f in os.listdir(dir_fotos_input)
                         if f.lower().endswith(
                             ('.jpg','.jpeg','.png','.gif','.bmp','.webp'))])
                st.success(f"✅ Pasta configurada: {dir_fotos_input} "
                           f"({n} imagens)")
            else:
                st.error(f"❌ Pasta não encontrada: {dir_fotos_input}")
    with col_dir2:
        if st.button("📂 Abrir no Explorer", use_container_width=True,
                     key="btn_abrir_dir_fotos"):
            try:
                if platform.system() == 'Windows':
                    subprocess.Popen(
                        ['explorer', dir_fotos_input.replace('/', '\\')])
                    st.success("✅ Abrindo no Explorer...")
                else:
                    st.warning("⚠️ Só funciona no Windows local.")
            except Exception as ex:
                st.warning(f"⚠️ Erro: {ex}")

    if 'dir_fotos_manual' in st.session_state:
        st.info(f"📂 Pasta de fotos ativa: "
                f"**{st.session_state['dir_fotos_manual']}**")

st.markdown("---")

# ═══════════════════════════════════════════════════════════════════════════════
# Painel principal — aparece após dados carregados
# ═══════════════════════════════════════════════════════════════════════════════
csv_file = None
if 'csv_data' in st.session_state:
    csv_file = BytesIO(st.session_state['csv_data'].encode('utf-8'))
    fonte = ("ODK Central"
             if st.session_state.get('data_source') == 'odk'
             else "Upload Manual")
    st.info(f"📊 Dados carregados de: **{fonte}**")

    # ── Botões CSV + pastas ───────────────────────────────────────────────────
    st.subheader("📁 Arquivos Locais")
    col_csv1, col_csv2, col_csv3 = st.columns(3)

    with col_csv1:
        st.download_button(
            label="⬇️ Baixar Planilha CSV",
            data=st.session_state['csv_data'].encode('utf-8'),
            file_name="dados_odk.csv",
            mime="text/csv",
            use_container_width=True,
            key="btn_download_csv",
        )

    with col_csv2:
        if st.button("📂 Abrir C:/sepe", use_container_width=True,
                     key="btn_abrir_sepe"):
            try:
                if platform.system() == 'Windows':
                    os.makedirs('C:/sepe', exist_ok=True)
                    subprocess.Popen(['explorer', 'C:\\sepe'])
                    st.success("✅ Abrindo C:/sepe...")
                else:
                    st.warning("⚠️ Só funciona no Windows local.")
            except Exception as ex:
                st.warning(f"⚠️ Erro: {ex}")

    with col_csv3:
        if st.button("📂 Abrir C:/sepe/media", use_container_width=True,
                     key="btn_abrir_media"):
            try:
                if platform.system() == 'Windows':
                    os.makedirs('C:/sepe/media', exist_ok=True)
                    subprocess.Popen(['explorer', 'C:\\sepe\\media'])
                    st.success("✅ Abrindo C:/sepe/media...")
                else:
                    st.warning("⚠️ Só funciona no Windows local.")
            except Exception as ex:
                st.warning(f"⚠️ Erro: {ex}")

    st.markdown("---")

    # ── Download ZIP de imagens ───────────────────────────────────────────────
    if st.session_state.get('anexos_baixados'):
        st.subheader("📥 Download de Imagens")
        col_img1, col_img2 = st.columns(2)
        with col_img1:
            st.info(f"**{len(st.session_state['anexos_baixados'])} imagens** "
                    f"disponíveis")
        with col_img2:
            zip_buf = BytesIO()
            with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                for anexo in st.session_state['anexos_baixados']:
                    zf.writestr(anexo['nome_com_prefixo'], anexo['data'])
            zip_buf.seek(0)
            st.download_button(
                label="📦 Baixar Todas as Imagens (ZIP)",
                data=zip_buf,
                file_name="imagens_odk.zip",
                mime="application/zip",
                use_container_width=True,
                key="btn_download_zip_imagens",
            )
        st.markdown("---")

# ── Diretórios temporários ────────────────────────────────────────────────────
@st.cache_resource
def criar_diretorios_temp():
    base = tempfile.mkdtemp()
    dirs = {'base': base}
    for nome in ['xlsx', 'relatorios', 'media', 'modelo']:
        p = os.path.join(base, nome)
        os.makedirs(p, exist_ok=True)
        dirs[nome] = p
    return dirs

dirs = criar_diretorios_temp()

# ── Modelo do relatório ───────────────────────────────────────────────────────
st.subheader("📄 Modelo do Relatório")
modelo_file = st.file_uploader("Upload do modelo DOCX (formulario.docx)",
                                type=['docx'])

# Info sobre pasta de imagens ativa
st.subheader("📁 Pasta de Imagens para Relatórios")
dir_ativo = st.session_state.get('dir_fotos_manual', 'C:/sepe/media')
if os.path.isdir(dir_ativo):
    n_imgs = len([f for f in os.listdir(dir_ativo)
                  if f.lower().endswith(
                      ('.jpg','.jpeg','.png','.gif','.bmp','.webp'))])
    st.success(f"✅ Usando: **{dir_ativo}** ({n_imgs} imagens encontradas)")
else:
    st.info(f"📂 Pasta configurada: **{dir_ativo}** "
            f"(não encontrada localmente — app pode estar em servidor remoto)")

st.markdown("---")

# ═══════════════════════════════════════════════════════════════════════════════
# Funções de processamento
# ═══════════════════════════════════════════════════════════════════════════════

def imagem_valida(path: str) -> bool:
    """Valida com PIL se o arquivo é uma imagem real (evita UnrecognizedImageError)."""
    try:
        from PIL import Image as PILImage
        with PILImage.open(path) as img:
            img.verify()
        return True
    except Exception:
        return False


def processar_imagem(doc, valor_imagem, dirs):
    """
    Retorna InlineImage validado com PIL ou None.
    Ordem de busca:
      1. Pasta manual (dir_fotos_manual do session_state)
      2. C:/sepe/media/
      3. Pasta temporária ODK
      4. Pasta interna temporária
    Sem imagem → imagem padrão (xxx.jpg local ou download).
    """

    def inline(path, cm):
        if not os.path.exists(path):
            return None
        if not imagem_valida(path):
            print(f"Arquivo inválido (não é imagem): {path}")
            return None
        try:
            return InlineImage(doc, path, Cm(cm))
        except Exception as ex:
            print(f"InlineImage falhou ({path}): {ex}")
            return None

    def imagem_padrao():
        r = inline('C:/sepe/xxx.jpg', 3)
        if r:
            return r
        try:
            url = ("https://st2.depositphotos.com/12694644/47297/v/380/"
                   "depositphotos_472972706-stock-illustration-image-available"
                   "-sign-isolated-white.jpg")
            tmp = os.path.join(tempfile.gettempdir(), 'no_image_default.jpg')
            if not os.path.exists(tmp):
                resp = requests.get(url, timeout=10)
                resp.raise_for_status()
                ct = resp.headers.get('Content-Type', '')
                if 'image' not in ct.lower():
                    raise ValueError(f"Content-Type inválido: {ct}")
                Path(tmp).write_bytes(resp.content)
            r = inline(tmp, 3)
            if r:
                return r
        except Exception as ex:
            print(f"Imagem padrão internet falhou: {ex}")
        return None

    if not valor_imagem or str(valor_imagem).strip() == '':
        return imagem_padrao()

    nome = str(valor_imagem).strip()

    candidatos = []
    dir_manual = st.session_state.get('dir_fotos_manual', '')
    if dir_manual:
        candidatos.append(os.path.join(dir_manual, nome))
    candidatos += [
        f'C:/sepe/media/{nome}',
        os.path.join(tempfile.gettempdir(), 'odk_media', nome),
        os.path.join(dirs.get('media', ''), nome),
    ]

    for path in candidatos:
        r = inline(path, 12)
        if r:
            return r

    print(f"Imagem '{nome}' não encontrada. Usando padrão.")
    return imagem_padrao()


def converter_csv_para_xlsx(csv_file, xlsx_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'dados_vistoria'
    linhas = csv.reader(csv_file.getvalue().decode('utf-8').splitlines())
    for row_idx, row in enumerate(linhas, start=1):
        ws.cell(row=row_idx, column=1, value=row_idx)
        for col_idx, val in enumerate(row, start=2):
            ws.cell(row=row_idx, column=col_idx, value=val)
    wb.save(xlsx_path)
    return xlsx_path


def processar_relatorios(xlsx_path, modelo_path, dirs,
                          indices_selecionados=None):
    wb    = openpyxl.load_workbook(xlsx_path)
    rows  = list(wb['dados_vistoria'].values)

    if indices_selecionados:
        dados = [rows[0]] + [rows[i] for i in indices_selecionados
                             if i < len(rows)]
    else:
        dados = rows

    total  = len(dados) - 1
    prog   = st.progress(0)
    status = st.empty()
    gerados = []

    for idx, v in enumerate(dados[1:], 1):
        status.text(f"Processando {idx}/{total}: {v[0]}")
        prog.progress(idx / total)

        doc  = DocxTemplate(modelo_path)
        img1 = processar_imagem(doc, v[18], dirs)
        img2 = processar_imagem(doc, v[19], dirs)
        img3 = processar_imagem(doc, v[20], dirs)
        img4 = processar_imagem(doc, v[21], dirs)
        img5 = processar_imagem(doc, v[22], dirs)

        data_fmt = v[2]
        if v[2] and isinstance(v[2], str):
            try:
                from datetime import datetime
                dt = (datetime.fromisoformat(v[2].replace('Z', '+00:00'))
                      if 'T' in v[2]
                      else datetime.strptime(v[2], '%Y-%m-%d'))
                data_fmt = dt.strftime('%d-%m-%Y')
            except Exception:
                pass

        doc.render({
            'possui_placa':  v[9],
            'plano_trabalho': v[10],
            'relatorio':     v[0],
            'id_proj':       v[5],
            'id_tipo_rel':   v[8],
            'meta':          v[25],
            'data':          data_fmt,
            'processo_sei':  v[6],
            'cidade':        v[12],
            'responsavel':   v[27],
            'lat':           v[13],
            'long':          v[14],
            'observacao':    v[23],
            'tipo_proj':     v[17],
            'imagem_1':      img1,
            'imagem_2':      img2,
            'imagem_3':      img3,
            'imagem_4':      img4,
            'imagem_5':      img5,
        })

        out = os.path.join(dirs['relatorios'], f"{v[0]}.docx")
        doc.save(out)
        gerados.append(out)

    prog.empty()
    status.empty()
    return gerados


def criar_zip(arquivos, zip_path):
    with zipfile.ZipFile(zip_path, 'w') as zf:
        for arq in arquivos:
            zf.write(arq, os.path.basename(arq))


# ═══════════════════════════════════════════════════════════════════════════════
# Preview / seleção de relatórios
# ═══════════════════════════════════════════════════════════════════════════════
indices_selecionados = []

if csv_file is not None:
    st.subheader("📋 Visualizar e Selecionar Relatórios")
    csv_file.seek(0)
    lines = csv_file.read().decode('utf-8').strip().split('\n')

    df = pd.DataFrame()
    header = []

    if len(lines) > 1:
        original_cols = next(csv.reader([lines[0]]))
        seen, unique_cols = {}, []
        for col in original_cols:
            col = col.strip()
            if col not in seen:
                seen[col] = 0
                unique_cols.append(col)
            else:
                seen[col] += 1
                unique_cols.append(f"{col}_dup{seen[col]}")

        data_rows = [r for r in csv.reader(lines[1:]) if r]
        df = pd.DataFrame(data_rows, columns=unique_cols)

        cols_final = []
        for i, c in enumerate(df.columns):
            cols_final.append(
                f"{c}_idx{i}" if df.columns.tolist().count(c) > 1 else c)
        df.columns = cols_final
        df.insert(0, '#', range(1, len(df) + 1))
        header = df.columns.tolist()

        if len(header) != len(set(header)):
            st.error(f"🔴 Duplicatas: "
                     f"{[h for h in header if header.count(h) > 1]}")
            st.stop()

        for col in [c for c in df.columns
                    if 'SubmissionDate' in c and not c[-1].isdigit()]:
            try:
                df[col] = (pd.to_datetime(df[col], errors='coerce')
                           .dt.strftime('%d-%m-%Y'))
            except Exception:
                pass

    if len(df) > 0:
        st.info(f"📊 Total disponível: **{len(df)}** relatórios")
        col_sel1, col_sel2 = st.columns([1, 3])

        def colunas_display(header, df, tipo_proj=False):
            cols = ['#']
            buscas = [
                (['N_mero_ID', 'Numero_ID', 'details-N'], 1),
                (['Tipo_Relat', 'Tipo_Relatorio'],         None),
                (['SubmissionDate'],                       3),
                (['cidade', 'regiao'],                     7),
                (['processo', 'sei'],                      6),
            ]
            for padroes, fallback in buscas:
                achados = [c for c in header if any(p in c for p in padroes)]
                if achados:
                    cols.append(achados[0])
                elif fallback and len(header) > fallback:
                    cols.append(header[fallback])
            if tipo_proj:
                tp = [c for c in header
                      if 'tipo' in c.lower() and 'proj' in c.lower()]
                if tp:
                    cols.append(tp[0])
                elif len(header) > 12:
                    cols.append(header[12])
            vistos = []
            for c in cols:
                if c not in vistos and c in df.columns:
                    vistos.append(c)
            return vistos

        with col_sel1:
            selecao_tipo = st.radio(
                "Tipo de seleção:",
                ["Todos os relatórios", "Selecionar específicos"],
                key="tipo_selecao")

        with col_sel2:
            if selecao_tipo == "Selecionar específicos":
                df_disp = (df[colunas_display(header, df)]
                           .copy().reset_index(drop=True))
                st.dataframe(df_disp, use_container_width=True, height=400)
                nums = st.text_input(
                    "Números dos relatórios (ex: 1, 3, 5-10):",
                    placeholder="1, 3, 5-10")
                if nums:
                    try:
                        for parte in nums.split(','):
                            parte = parte.strip()
                            if '-' in parte:
                                a, b = map(int, parte.split('-'))
                                indices_selecionados.extend(range(a, b+1))
                            else:
                                indices_selecionados.append(int(parte))
                        indices_selecionados = sorted(set(indices_selecionados))
                        st.success(
                            f"✅ {len(indices_selecionados)} selecionados: "
                            f"{', '.join(map(str, indices_selecionados))}")
                    except Exception:
                        st.error("❌ Formato inválido.")
            else:
                df_disp = (df[colunas_display(header, df, tipo_proj=True)]
                           .copy().reset_index(drop=True))
                st.dataframe(df_disp, use_container_width=True, height=400)
                st.caption(f"📊 Todos os {len(df)} relatórios serão gerados")
                indices_selecionados = list(range(1, len(df)+1))
    else:
        st.warning("⚠️ CSV vazio.")

st.markdown("---")

# ═══════════════════════════════════════════════════════════════════════════════
# Gerar relatórios
# ═══════════════════════════════════════════════════════════════════════════════
habilitado = (csv_file is not None
              and modelo_file is not None
              and len(indices_selecionados) > 0)

if csv_file is not None and modelo_file is not None and not habilitado:
    st.warning("⚠️ Selecione ao menos um relatório.")

if st.button("🚀 Gerar Relatórios", type="primary",
             use_container_width=True, disabled=not habilitado):
    try:
        with st.spinner("Processando..."):
            modelo_path = os.path.join(dirs['modelo'], 'formulario.docx')
            Path(modelo_path).write_bytes(modelo_file.getbuffer())

            st.info("Convertendo CSV para XLSX...")
            xlsx_path = os.path.join(dirs['xlsx'], 'dados.xlsx')
            converter_csv_para_xlsx(csv_file, xlsx_path)

            st.info("Gerando relatórios...")
            relatorios = processar_relatorios(
                xlsx_path, modelo_path, dirs, indices_selecionados)

            zip_path = os.path.join(dirs['base'], 'relatorios.zip')
            criar_zip(relatorios, zip_path)

            st.success(f"✅ {len(relatorios)} relatórios gerados!")
            st.download_button(
                label="📥 Download ZIP com Relatórios DOCX",
                data=Path(zip_path).read_bytes(),
                file_name="relatorios_vistoria.zip",
                mime="application/zip",
                use_container_width=True,
                key="btn_download_relatorios",
            )
    except Exception as e:
        st.error(f"❌ Erro: {e}")
        st.exception(e)

st.markdown("---")
st.caption("Desenvolvido para SEPE - Sistema de Geração de Relatórios de "
           "Vistoria - versão 2.0")
